import openpyxl as xl

# Opening the source file.
source_file = r'C:Enter\file\path\here'
wb = xl.load_workbook(source_file)
ws = wb.worksheets[0]

# Opening the destination file.
destination_file = r'C:Enter\file\path\here'
wb2 = r'C:Enter\file\path\here'
ws2 = wb.active

# Calculating the total number of rows & columns in the source file.
total_row = ws.max_row
total_column = ws.max_column

# Copying the source excel file to the destionation excel file.
for i in range(1, total_row + 1):
    for j in range(1, total_column + 1):
        # Reading the cell value from the source excel file
        cell_value = ws.cell(row = i, column = j)

        # Writing the read value to the destination excel file.
        ws2.cell(row = i, column = j).value = cell_value.value

# Saving the destionation file.
wb2.save(str(destination_file))

# https://www.geeksforgeeks.org/python-how-to-copy-data-from-one-excel-sheet-to-another/
